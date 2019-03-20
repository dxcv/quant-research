
# ----------------更新eod数据（最好在13：45之后更新，有时候早一点）----------------------------------------------------------------------------
from tools.data import update
import const

endTime = const.TODAY

update.trade_dates(endTime=endTime)
update.eod_ashare(endTime=endTime)
update.eod_indexes(endTime=endTime)  # 881001只到2017-11-24
update.eod_induIndex_citics(endTime=endTime)
update.eod_index_wind(endTime=endTime)
update.eod_citics_industry_type(endTime=endTime)  # @ todo 出问题了


# ---------使用的模板----------------------------------------------------------------------------------------------------
import pandas as pd
from WindPy import w
from dateutil.parser import parse
from tools.data import fetch
from openpyxl import load_workbook



# ---------要调的参数----------------------------------------------------------------------------------------------------
day = const.TODAY
tradeDate = fetch.trade_dates(startTime='2019-01-01',endTime=day)
lastTradeDate = tradeDate.iat[-2,0].strftime('%Y-%m-%d')
targetFp = r'C:\Users\huangshengshan\Desktop\新日报模板.xlsx'
# 主线先把IT是数据拷贝到sheet2
itDataSource = r'C:\Users\huangshengshan\Desktop\权益类组合每日交易概览.xlsx'


someIndex = fetch.index_list(['000001.SH', '000016.SH', '000300.SH', '000852.SH', '399006.SZ',],
                             startTime='2019-01-01', endTime=day)

if someIndex.index.get_level_values(0)[-1].strftime('%Y-%m-%d') != const.TODAY:
    raise ValueError('今天的行情数据还没更新，请检查!')

# ----------------------------------------------------------------------------------------------------------------------
def _fenzu(x):
    try:
        return portofolioGroupSer[x]
    except KeyError:
        return '其他'


def _del_other(x):
    # @todo 这个会出问题，注意修改，主要是排除优先股和转债
    try:
        rst = f'{int(x):0>6}'
        if rst[0] == '1' or rst[:2] == '36' or rst == '600928':
            return 'other'
        return rst
    except:
        return 'other'


def _get_citics_classification(x):
    try:
        return dfCiticsContent[x]
    except:
        return '未上市'


# ------------基础的表格-------------------------------------------------------------------------------------------------
# 基金的风格分类表
portofolioGroup = pd.read_excel(r'\\ia\temp\jyb\HSS\日报基础数据\3权益成交回报统计汇总模板20180316版.xlsx',
                                sheet_name='风格小组',
                                usecols=[0,1,2,3,4,5])
portofolioGroup.drop_duplicates(inplace=True)
portofolioGroup['实际管理人'].fillna(method='bfill',inplace=True)
portofolioGroup.set_index('O32代码', inplace=True)
portofolioGroupSer = portofolioGroup['风格组']


# 中信行业对应表
dfCiticsContent = fetch.general_all('eod_ashare_industry_type_citics', startTime=day
                                    ).reset_index(level=0,drop=True)['INDUSTRY_CITIC_1']
dfCiticsContent.index = dfCiticsContent.reset_index()['CODE'].apply(lambda x: x[:6])

dfciticsUnclassified = pd.read_excel(r'\\ia\temp\jyb\HSS\日报基础数据\中信行业未纳入.xlsx', index_col=0)
dfciticsUnclassified.index = dfciticsUnclassified.reset_index()['CODE'].apply(lambda x: f'{int(x):>06}')
dfCiticsContent[dfCiticsContent == ''] = dfciticsUnclassified['INDUSTRY_CITIC_1']


#  维护未纳入个股中信行业，
if not dfCiticsContent[dfCiticsContent == ''].empty:
    print(dfCiticsContent[dfCiticsContent == ''].index)
    raise ValueError('有中信行业未纳入个股，请按照上面的代码填写为纳入个股的行业')


# 读取公司成交数据
tradeData = pd.read_excel(r'\\ia\temp\jyb\HSS\日报数据源\成交数据{day}.xls'.format(day=day))[:-1]
tradeData['发生日期'] = tradeData['发生日期'].apply(lambda x: parse(x))
tradeData = tradeData[(tradeData['交易市场'] == '深交所A') | (tradeData['交易市场'] == '上交所A')]  # 排除港股
tradeData['证券代码'] = tradeData['证券代码'].apply(lambda x: f'{int(x):0>6}')
tradeData['基金编号'] = tradeData['基金编号'].apply(int)
tradeData['中信一级行业'] = tradeData['证券代码'].apply(lambda x: dfCiticsContent[x])
tradeData['分组'] = tradeData['基金编号'].apply(_fenzu)


# 全市场成交金额
w.start()
windData = w.wsd("881001.WI", "amt", day, day, "")
w.stop()


# 个股持仓明细，计算配置状态
chicangDetail = pd.read_excel(r'\\ia\temp\jyb\HSS\日报数据源\组合证券{day}.xls'.format(day=day))[:-1]
chicangDetail['日期'] = chicangDetail['日期'].apply(lambda x: parse(x))
chicangDetail = chicangDetail[(chicangDetail['交易市场'] == '深交所A') | (chicangDetail['交易市场'] == '上交所A')]
chicangDetail['证券代码'] = chicangDetail['证券代码'].apply(_del_other)
chicangDetail = chicangDetail[chicangDetail['证券代码'] != 'other']  # 排除状态优先股
chicangDetail['基金编号'] = chicangDetail['基金编号'].apply(int)
chicangDetail['中信一级行业'] = chicangDetail['证券代码'].apply(_get_citics_classification)
chicangDetail['分组'] = chicangDetail['基金编号'].apply(_fenzu)
chicangDetail = chicangDetail[chicangDetail['分组'] != '其他']  # 不计算其他分组的情况


# hs300权重
hs300weight = pd.read_excel(r'\\ia\temp\jyb\HSS\日报基础数据\hs300权重.xlsx',)  # 这个表格冲wind下载的不能直接读 @todo
hs300weight['代码'] = hs300weight['代码'].apply(lambda x:x[:6])
hs300weight['中信一级行业'] = hs300weight['代码'].apply(lambda x: dfCiticsContent[x])
hs300weightSer = hs300weight.groupby('中信一级行业')['权重（%）'].sum()


# 中信行业指数的涨跌幅
dfIndu = fetch.index_industry_list('all', citicslevel=1, startTime=day)
dfIndu['S_INFO_NAME'] = dfIndu['S_INFO_NAME'].apply(lambda x: x[:-4])


# --------一、市场回顾checked--------------------------------------------------------------------------------------------
# tradeDate = fetch.trade_dates(startTime='2019-01-01',endTime=day)
# someIndex = fetch.index_list(['000001.SH', '000016.SH', '000300.SH', '000852.SH', '399006.SZ',],
#                              startTime='2019-01-01',endTime=day)

date5before = tradeDate['DATETIME'].iat[-5]
date20before = tradeDate['DATETIME'].iat[-20]
dateFirst = tradeDate['DATETIME'].iat[0]

shichanghuiguRst = pd.DataFrame(index=['上证综指', '上证50', '沪深300', '中证1000', '创业板指'],
                                columns=['涨跌幅(%)', '前5日涨跌幅(%)', '前20日涨跌幅(%)', '年初至今(%)'])
shichanghuiguRst['涨跌幅(%)'] = someIndex.loc[day,'PCTCHANGE'].values
shichanghuiguRst['前5日涨跌幅(%)'] \
    = ((someIndex.loc[parse(day), 'CLOSE'] / someIndex.loc[date5before, 'PRECLOSE'] - 1) * 100).values
shichanghuiguRst['前20日涨跌幅(%)'] \
    = ((someIndex.loc[parse(day), 'CLOSE'] / someIndex.loc[date20before, 'PRECLOSE'] - 1) * 100).values
shichanghuiguRst['年初至今(%)'] \
    = ((someIndex.loc[parse(day), 'CLOSE'] / someIndex.loc[dateFirst, 'PRECLOSE'] - 1) * 100).values


# -------二、公司成交情况checked------------------------------------------------------------------------------------------
gongsiTradeRst = pd.DataFrame(index=[0,], columns=['买入成交金额', '卖出成交金额', '总成交金额', '市场成交占比'] )
gongsiTradeRst.at[0, '买入成交金额'] = tradeData.loc[tradeData['委托方向'] == '买入', '成交金额'].sum() / 1e8
gongsiTradeRst.at[0, '卖出成交金额'] = tradeData.loc[tradeData['委托方向'] == '卖出', '成交金额'].sum() / 1e8
gongsiTradeRst.at[0, '总成交金额'] = gongsiTradeRst.at[0, '买入成交金额'] + gongsiTradeRst.at[0, '卖出成交金额']
gongsiTradeRst.at[0, '市场成交占比'] = gongsiTradeRst.at[0, '总成交金额'] / windData.Data[0][0] * 1e8


# ---------二、3集中交易的行业对应，从这里拿sheet2的数据checked-------------------------------------------------------------
wb = load_workbook(itDataSource)
sh = wb["Sheet2"]

_equityPoi = sh['B68'].value
_top30TodayPoi = sh['B90'].value

jizhongTradeRst = pd.DataFrame(index=range(10), columns=['代码', '行业'])

i = 0
for cell in sh['A114':'A118']:
    jizhongTradeRst.at[i, '代码'] = f'{cell[0].value:0>6}'
    i += 1

for cell in sh['A121':'A125']:
    jizhongTradeRst.at[i, '代码'] = f'{cell[0].value:0>6}'
    i += 1

wb.close()
del wb, sh, cell
jizhongTradeRst['行业'] = jizhongTradeRst['代码'].apply(lambda x: dfCiticsContent[x])


# ---------三1、行业交易概况---------------------------------------------------------------------------------------------
top3 = list(dfIndu.sort_values(by='PCTCHANGE', ascending=False)['S_INFO_NAME'][:3])
top3Rst = pd.DataFrame(index=[0,1,2], columns=['涨幅前三行业',
                                               '买入成交金额',
                                               '买入金额占全天公司买入金额比例(%)',
                                               '行业配置状态(是否超配)',
                                               '金额前三大个股'])
top3Rst['涨幅前三行业'] = top3

for i in [0,1,2]:
    top3Rst.at[i, '买入成交金额'] = tradeData.loc[
        (tradeData['中信一级行业'] == top3[i])
        & (tradeData['委托方向'] == '买入')
        & (tradeData['发生日期'] == parse(day)),
        '成交金额'
    ].sum() / 1e8

    top3Rst.at[i, '买入金额占全天公司买入金额比例(%)'] \
        = top3Rst.at[i, '买入成交金额'] * 1e8 \
          / tradeData.loc[
              (tradeData['委托方向'] == '买入')
              & (tradeData['发生日期'] == parse(day)),
              '成交金额'
          ].sum()

    top3Rst.at[i, '金额前三大个股'] = list(
        tradeData[
        (tradeData['中信一级行业'] == top3[i])
        & (tradeData['委托方向'] == '买入')
        & (tradeData['发生日期'] == parse(day))
    ].groupby('证券名称')['成交金额'].sum().sort_values(ascending=False)[:3].index
    )
    top3Rst.at[i, '金额前三大个股'] = '\n'.join(top3Rst.at[i, '金额前三大个股'])

    _ratio = chicangDetail.loc[(chicangDetail['中信一级行业'] == top3[i]), '持仓'].sum() / chicangDetail['持仓'].sum()
    top3Rst.at[i, '行业配置状态(是否超配)'] = f'{_ratio * 100:.2f}%\n超配' if (_ratio - (
                hs300weightSer[top3[i]] / 100)) > 0 else f'{_ratio * 100:.2f}%\n低配'

# 跌幅靠前的
bottom3 = list(dfIndu.sort_values(by='PCTCHANGE', ascending=False)['S_INFO_NAME'][-3:])
bottom3Rst = pd.DataFrame(index=[0,1,2], columns=['跌幅前三行业',
                                                  '卖出成交金额',
                                                  '卖出金额占全天公司卖出金额比例(%)',
                                                  '行业配置状态(是否超配)',
                                                  '金额前三大个股'])
bottom3Rst['跌幅前三行业'] = bottom3

for i in [0,1,2]:
    bottom3Rst.at[i, '卖出成交金额'] = tradeData.loc[
        (tradeData['中信一级行业'] == bottom3[i])
        & (tradeData['委托方向'] == '卖出')
        & (tradeData['发生日期'] == parse(day)), '成交金额'
    ].sum() / 1e8

    bottom3Rst.at[i, '卖出金额占全天公司卖出金额比例(%)']\
        = bottom3Rst.at[i, '卖出成交金额'] * 1e8\
          / tradeData.loc[
              (tradeData['委托方向'] == '卖出') & (tradeData['发生日期'] == parse(day)), '成交金额'
          ].sum()

    bottom3Rst.at[i, '金额前三大个股'] = list(
        tradeData[
        (tradeData['中信一级行业'] == bottom3[i])
        & (tradeData['委托方向'] == '卖出')
        & (tradeData['发生日期'] == parse(day))
    ].groupby('证券名称')['成交金额'].sum().sort_values(ascending=False)[:3].index
    )

    bottom3Rst.at[i, '金额前三大个股'] = '\n'.join(bottom3Rst.at[i, '金额前三大个股'])
    _ratio = chicangDetail.loc[(chicangDetail['中信一级行业'] == bottom3[i]), '持仓'].sum() / chicangDetail['持仓'].sum()
    bottom3Rst.at[i, '行业配置状态(是否超配)'] = f'{_ratio * 100:.2f}%\n超配' if (_ratio - (
                hs300weightSer[bottom3[i]] / 100)) > 0 else f'{_ratio * 100:.2f}%\n低配'

# 买卖金额最大的行业
buyTop3Rst = pd.DataFrame(index=[0,1,2], columns=['买入金额前三行业',
                                                  '买入成交金额',
                                                  '买入金额占全天公司买入金额比例(%)',
                                                  '行业配置状态(是否超配)',
                                                  '金额前三大个股'])
buyTop3Rst['买入金额前三行业'] = tradeData.loc[
                             (tradeData['委托方向'] == '买入')
                            & (tradeData['发生日期'] == parse(day))
            ].groupby('中信一级行业')['成交金额'].sum().sort_values(ascending=False).index[:3]

for i in [0,1,2]:
    buyTop3Rst.at[i, '买入成交金额'] = tradeData.loc[
        (tradeData['中信一级行业'] == buyTop3Rst.at[i, '买入金额前三行业'])
        & (tradeData['委托方向'] == '买入')
        & (tradeData['发生日期'] == parse(day)), '成交金额'
    ].sum() / 1e8

    buyTop3Rst.at[i, '买入金额占全天公司买入金额比例(%)'] \
        = buyTop3Rst.at[i, '买入成交金额'] * 1e8 \
          / tradeData.loc[
              (tradeData['委托方向'] == '买入') & (tradeData['发生日期'] == parse(day)), '成交金额'
          ].sum()

    buyTop3Rst.at[i, '金额前三大个股'] = list(
        tradeData[
        (tradeData['中信一级行业'] == buyTop3Rst.at[i, '买入金额前三行业'])
        & (tradeData['委托方向'] == '买入')
        & (tradeData['发生日期'] == parse(day))
    ].groupby('证券名称')['成交金额'].sum().sort_values(ascending=False)[:3].index
    )

    buyTop3Rst.at[i, '金额前三大个股'] = '\n'.join(buyTop3Rst.at[i, '金额前三大个股'])
    _ratio = chicangDetail.loc[(chicangDetail['中信一级行业'] == buyTop3Rst.at[i, '买入金额前三行业']), '持仓'].sum() / chicangDetail['持仓'].sum()
    buyTop3Rst.at[i, '行业配置状态(是否超配)'] = f'{_ratio * 100:.2f}%\n超配' if (_ratio - (
                hs300weightSer[buyTop3Rst.at[i, '买入金额前三行业']] / 100)) > 0 else f'{_ratio * 100:.2f}%\n低配'

sellTop3Rst = pd.DataFrame(index=[0,1,2], columns=['卖出金额前三行业',
                                                   '卖出成交金额',
                                                   '卖出金额占全天公司卖出金额比例(%)',
                                                   '行业配置状态(是否超配)',
                                                   '金额前三大个股'])
sellTop3Rst['卖出金额前三行业'] = tradeData.loc[(tradeData['委托方向'] == '卖出')
        & (tradeData['发生日期'] == parse(day))
            ].groupby('中信一级行业')['成交金额'].sum().sort_values(ascending=False).index[:3]

for i in [0,1,2]:
    sellTop3Rst.at[i, '卖出成交金额'] = tradeData.loc[
        (tradeData['中信一级行业'] == sellTop3Rst.at[i, '卖出金额前三行业'])
        & (tradeData['委托方向'] == '卖出')
        & (tradeData['发生日期'] == parse(day)), '成交金额'
    ].sum() / 1e8

    sellTop3Rst.at[i, '卖出金额占全天公司卖出金额比例(%)'] \
        = sellTop3Rst.at[i, '卖出成交金额'] * 1e8 \
          / tradeData.loc[
              (tradeData['委托方向'] == '卖出') & (tradeData['发生日期'] == parse(day)), '成交金额'
          ].sum()
    sellTop3Rst.at[i, '金额前三大个股'] = list(
        tradeData[
        (tradeData['中信一级行业'] == sellTop3Rst.at[i, '卖出金额前三行业'])
        & (tradeData['委托方向'] == '卖出')
        & (tradeData['发生日期'] == parse(day))
    ].groupby('证券名称')['成交金额'].sum().sort_values(ascending=False)[:3].index
    )
    sellTop3Rst.at[i, '金额前三大个股'] = '\n'.join(sellTop3Rst.at[i, '金额前三大个股'])
    # sellTop3Rst.at[i, '行业配置状态(是否超配)'] \
    #     = (chicangDetail.loc[(chicangDetail['中信一级行业'] == sellTop3Rst.at[i, '卖出金额前三行业']), '持仓'].sum() / chicangDetail['持仓'].sum()) - (hs300weightSer[sellTop3Rst.at[i, '卖出金额前三行业']]/100)
    # sellTop3Rst.at[i, '行业配置状态(是否超配)'] = '是' if sellTop3Rst.at[i, '行业配置状态(是否超配)'] > 0 else '否'

    _ratio = chicangDetail.loc[(chicangDetail['中信一级行业'] == sellTop3Rst.at[i, '卖出金额前三行业']), '持仓'].sum() / chicangDetail['持仓'].sum()
    sellTop3Rst.at[i, '行业配置状态(是否超配)'] = f'{_ratio * 100:.2f}%\n超配' if (_ratio - (
                hs300weightSer[sellTop3Rst.at[i, '卖出金额前三行业']] / 100)) > 0 else f'{_ratio * 100:.2f}%\n低配'


# 三、2风格小组交易概况
fenzuPoi = pd.read_excel(r'\\ia\temp\jyb\HSS\日报数据源\基金资产{day}.xls'.format(day=day))[:-1]
fenzuPoi = fenzuPoi[fenzuPoi['净值'] > 0]
fenzuPoi['基金编号'] = fenzuPoi['基金编号'].apply(int)
fenzuPoi['分组'] = fenzuPoi['基金编号'].apply(_fenzu)
fenzuPoi = fenzuPoi[fenzuPoi['分组'] != '其他']

lastDatefenzuPoi = pd.read_excel(r'\\ia\temp\jyb\HSS\日报数据源\基金资产{day}.xls'.format(day=lastTradeDate))[:-1]
lastDatefenzuPoi = lastDatefenzuPoi[lastDatefenzuPoi['净值'] > 0]
lastDatefenzuPoi['基金编号'] = lastDatefenzuPoi['基金编号'].apply(int)
lastDatefenzuPoi['分组'] = lastDatefenzuPoi['基金编号'].apply(_fenzu)
lastDatefenzuPoi = lastDatefenzuPoi[lastDatefenzuPoi['分组'] != '其他']

fenzuPoiRst = pd.DataFrame(fenzuPoi.groupby('分组')['股票资产'].sum() / fenzuPoi.groupby('分组')['净值'].sum() * 100)
fenzuPoiRst['lastDate'] = lastDatefenzuPoi.groupby('分组')['股票资产'].sum() / lastDatefenzuPoi.groupby('分组')['净值'].sum() * 100
fenzuPoiRst['change'] = fenzuPoiRst[0] - fenzuPoiRst['lastDate']
fenzuPoiRst['当日情况'] = ''

wb = load_workbook(itDataSource)
sh = wb["Sheet2"]

fenzuPoiRst.at['GARP组', '当日情况'] = sh['A143'].value[:2] + f": {sh['B143'].value/1e8:.2f}\n" \
                                  + sh['A144'].value[:2] + f": {sh['B144'].value/1e8:.2f}\n" \
                                  + '净买入' + f": {-sh['B145'].value/1e8:.2f}"

fenzuPoiRst.at['成长组', '当日情况'] = sh['A160'].value[:2] + f": {sh['B160'].value/1e8:.2f}\n" \
                                  + sh['A161'].value[:2] + f": {sh['B161'].value/1e8:.2f}\n" \
                                  + '净买入' + f": {-sh['B162'].value/1e8:.2f}"

fenzuPoiRst.at['价值组', '当日情况'] = sh['A177'].value[:2] + f": {sh['B177'].value/1e8:.2f}\n" \
                                  + sh['A178'].value[:2] + f": {sh['B178'].value/1e8:.2f}\n" \
                                  + '净买入' + f": {-sh['B179'].value/1e8:.2f}"

fenzuPoiRst.at['绝对收益组', '当日情况'] = sh['A194'].value[:2] + f": {sh['B194'].value/1e8:.2f}\n" \
                                  + sh['A195'].value[:2] + f": {sh['B195'].value/1e8:.2f}\n" \
                                  + '净买入' + f": {-sh['B196'].value/1e8:.2f}"

fenzuPoiRst.at['主题组', '当日情况'] = sh['A211'].value[:2] + f": {sh['B211'].value/1e8:.2f}\n" \
                                  + sh['A212'].value[:2] + f": {sh['B212'].value/1e8:.2f}\n" \
                                  + '净买入' + f": {-sh['B213'].value/1e8:.2f}"

fenzuPoiRst.at['事业部', '当日情况'] = sh['A228'].value[:2] + f": {sh['B228'].value/1e8:.2f}\n" \
                                  + sh['A229'].value[:2] + f": {sh['B229'].value/1e8:.2f}\n" \
                                  + '净买入' + f": {-sh['B230'].value/1e8:.2f}"

wb.close()
del wb, sh

with open(r'\\ia\temp\jyb\HSS\日报数据源\30%lastDate.txt', 'r') as f:
    _top30LastDatePoi = float(f.readline())

with open(r'\\ia\temp\jyb\HSS\日报数据源\30%lastDate.txt', 'w') as f:
    f.write(f'{_top30TodayPoi}')

# 30%组合概览
with open(r'\\ia\temp\jyb\LQ\前30%.txt', 'r') as f:
    top30pct = f.read()

top30pct = [int(numStr) for numStr in top30pct.split(',')]
tradeData['前30%'] = tradeData['基金编号'].apply(lambda x: x in top30pct)

_top30buy = tradeData.loc[(tradeData['前30%']) & (tradeData['委托方向'] == '买入'), '成交金额'].sum() /1e8
_top30sell = tradeData.loc[(tradeData['前30%']) & (tradeData['委托方向'] == '卖出'), '成交金额'].sum() /1e8
_top30net = - tradeData.loc[(tradeData['前30%']), '发生金额(全价)'].sum() /1e8
top30pctRst = pd.DataFrame(index=[0,], columns=['前30%交易概况', '仓位变动'])
top30pctRst.at[0, '前30%交易概况'] = f'买入: {_top30buy:.2f}\n卖出: {_top30sell:.2f}\n净买入: {_top30net:.2f}'
top30pctRst.at[0, '仓位变动'] = (_top30TodayPoi - _top30LastDatePoi) * 100
del _top30buy, _top30sell, _top30net, top30pct


# ---二1和二2，历史仓位和成交金额的图--------------------------------------------------------------------------------------
dfHistoryPoi = pd.read_csv(r'\\ia\temp\jyb\HSS\日报数据源\历史仓位.csv', engine='python', encoding='gbk')
if parse(dfHistoryPoi['日期'].iat[-1]).strftime('%Y-%m-%d') != lastTradeDate:
    raise ValueError('历史仓位数据有误，最近一个不是昨天')
dfHistoryPoi.set_value(index=len(dfHistoryPoi), col=['日期', '沪深300', '权益类组合整体仓位'],
                       value = [day.replace('-', '/').replace('/0', '/'),
                                someIndex.at[(parse(day), '000300.SH'), 'CLOSE'],
                                _equityPoi
                                ])
dfHistoryPoi.to_csv(r'\\ia\temp\jyb\HSS\日报数据源\历史仓位.csv', encoding='gbk',index=False)

# 历史买卖和占比
dftradeRatio = pd.read_csv(r'\\ia\temp\jyb\HSS\日报数据源\历史成交占比.csv', engine='python', encoding='gbk')
if parse(dftradeRatio['日期'].iat[-1]).strftime('%Y-%m-%d') != lastTradeDate:
    raise ValueError('历史成交占比，最近一个不是昨天')
dftradeRatio.set_value(index=len(dftradeRatio), col=['日期', '买入成交金额', '卖出成交金额', '市场成交占比'],
                       value = [day.replace('-', '/').replace('/0', '/'),
                                gongsiTradeRst.at[0, '买入成交金额'],
                                gongsiTradeRst.at[0, '卖出成交金额'],
                                gongsiTradeRst.at[0, '市场成交占比']
                                ])
dftradeRatio.to_csv(r'\\ia\temp\jyb\HSS\日报数据源\历史成交占比.csv', encoding='gbk',index=False)
# ----------------------------------------------------------------------------------------------------------------------


# ------写入excel，写入老的模板，以免对格式造成干扰-------------------------------------------------------------------------
book = load_workbook(targetFp)
writer = pd.ExcelWriter(targetFp,engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

shichanghuiguRst.to_excel(writer, sheet_name='sheet4', startrow=0, startcol=0)

gongsiTradeRst.to_excel(writer, sheet_name='sheet4', index=False, startrow=6, startcol=0)

jizhongTradeRst.to_excel(writer, sheet_name='sheet4', index=False, startrow=8, startcol=0)

fenzuPoiRst.to_excel(writer,sheet_name='sheet4',startrow=19, startcol=0)
top30pctRst.to_excel(writer,sheet_name='sheet4',index=False, startrow=26, startcol=0)

top3Rst.to_excel(writer,sheet_name='sheet4', index=False,startrow=30, startcol=1)
bottom3Rst.to_excel(writer,sheet_name='sheet4', index=False,startrow=30, startcol=7)
buyTop3Rst.to_excel(writer,sheet_name='sheet4', index=False,startrow=35, startcol=1)
sellTop3Rst.to_excel(writer,sheet_name='sheet4',index=False,startrow=35, startcol=7)

dfHistoryPoi[-20:].to_excel(writer,sheet_name='sheet4',index=False, startrow=0, startcol=6)
dftradeRatio[-20:].to_excel(writer,sheet_name='sheet4',index=False, startrow=0, startcol=10)

writer.save()
writer.close()

